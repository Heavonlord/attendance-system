from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from flask import send_file
import io
from flask import Blueprint, render_template, redirect, url_for, flash, request
from flask_login import login_user, logout_user, login_required, current_user
from app import db
from app.models import User, Course, Attendance
from app.models import User, Course, Attendance, Enrollment  # Add Enrollment
from datetime import datetime

# Create blueprint
bp = Blueprint('main', __name__)

@bp.route('/')
def index():
    """Home page"""
    return render_template('index.html')

@bp.route('/login', methods=['GET', 'POST'])
def login():
    """Login page"""
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password):
            login_user(user)
            flash('Login successful!', 'success')
            
            # Redirect based on role
            if user.role == 'teacher':
                return redirect(url_for('main.teacher_dashboard'))
            else:
                return redirect(url_for('main.student_dashboard'))
        else:
            flash('Invalid username or password', 'danger')
    
    return render_template('login.html')

@bp.route('/logout')
@login_required
def logout():
    """Logout user"""
    logout_user()
    flash('You have been logged out', 'info')
    return redirect(url_for('main.index'))

@bp.route('/teacher/dashboard')
@login_required
def teacher_dashboard():
    """Teacher dashboard"""
    if current_user.role != 'teacher':
        flash('Access denied', 'danger')
        return redirect(url_for('main.index'))
    
    courses = Course.query.filter_by(teacher_id=current_user.id).all()
    return render_template('teacher_dashboard.html', courses=courses)

@bp.route('/student/dashboard')
@login_required
def student_dashboard():
    """Student dashboard"""
    if current_user.role != 'student':
        flash('Access denied', 'danger')
        return redirect(url_for('main.index'))
    
    # Get only enrolled courses
    enrollments = Enrollment.query.filter_by(student_id=current_user.id).all()
    enrolled_course_ids = [e.course_id for e in enrollments]
    
    # Get attendance records for enrolled courses only
    records = Attendance.query.filter(
        Attendance.student_id == current_user.id,
        Attendance.course_id.in_(enrolled_course_ids) if enrolled_course_ids else False
    ).all()
    
    # Calculate overall attendance percentage
    total = len(records)
    present = len([r for r in records if r.status == 'present'])
    percentage = (present / total * 100) if total > 0 else 0
    
    # Group records by course
    course_data = {}
    for record in records:
        if record.course_id not in course_data:
            course_data[record.course_id] = {
                'course': record.course,
                'total': 0,
                'present': 0,
                'records': []
            }
        course_data[record.course_id]['total'] += 1
        if record.status == 'present':
            course_data[record.course_id]['present'] += 1
        course_data[record.course_id]['records'].append(record)
    
    # Calculate percentages
    for course_id in course_data:
        data = course_data[course_id]
        data['percentage'] = round((data['present'] / data['total']) * 100, 2) if data['total'] > 0 else 0
    
    return render_template('student_dashboard.html', 
                         records=records,
                         percentage=round(percentage, 2),
                         course_data=course_data)
@bp.route('/teacher/mark-attendance/<int:course_id>', methods=['GET', 'POST'])
@login_required
def mark_attendance(course_id):
    """Mark attendance for a course"""
    if current_user.role != 'teacher':
        flash('Access denied', 'danger')
        return redirect(url_for('main.index'))
    
    course = Course.query.get_or_404(course_id)
    
    # Verify teacher owns this course
    if course.teacher_id != current_user.id:
        flash('You do not have access to this course', 'danger')
        return redirect(url_for('main.teacher_dashboard'))
    
    if request.method == 'POST':
        date_str = request.form.get('date')
        attendance_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        
    # Get all students
     # Get only enrolled students
    enrollments = Enrollment.query.filter_by(course_id=course_id).all()
    students = [e.student for e in enrollments]

    if not students:
        flash('No students enrolled in this course. Please enroll students first.', 'warning')
        return redirect(url_for('main.manage_enrollments', course_id=course_id))
        for student in students:
            status = request.form.get(f'status_{student.id}')
            if status:
                # Check if attendance already exists for this date
                existing = Attendance.query.filter_by(
                    student_id=student.id,
                    course_id=course_id,
                    date=attendance_date
                ).first()
                
                if existing:
                    existing.status = status
                else:
                    record = Attendance(
                        student_id=student.id,
                        course_id=course_id,
                        date=attendance_date,
                        status=status
                    )
                    db.session.add(record)
        
        db.session.commit()
        flash(f'Attendance marked for {len(students)} students', 'success')
        return redirect(url_for('main.teacher_dashboard'))
    
    # GET request - show form
    students = User.query.filter_by(role='student').all()
    today = datetime.now().date()
    
    return render_template('mark_attendance.html', 
                         course=course, 
                         students=students,
                         today=today)

@bp.route('/teacher/view-attendance/<int:course_id>')
@login_required
def view_course_attendance(course_id):
    """View all attendance records for a course"""
    if current_user.role != 'teacher':
        flash('Access denied', 'danger')
        return redirect(url_for('main.index'))
    
    course = Course.query.get_or_404(course_id)
    
    if course.teacher_id != current_user.id:
        flash('Access denied', 'danger')
        return redirect(url_for('main.teacher_dashboard'))
    
    # Get all attendance records for this course
    records = Attendance.query.filter_by(course_id=course_id).order_by(Attendance.date.desc()).all()
    
    # Group by student
    student_data = {}
    for record in records:
        if record.student_id not in student_data:
            student_data[record.student_id] = {
                'student': record.student,
                'records': [],
                'total': 0,
                'present': 0
            }
        student_data[record.student_id]['records'].append(record)
        student_data[record.student_id]['total'] += 1
        if record.status == 'present':
            student_data[record.student_id]['present'] += 1
    
    # Calculate percentages
    for student_id in student_data:
        data = student_data[student_id]
        if data['total'] > 0:
            data['percentage'] = round((data['present'] / data['total']) * 100, 2)
        else:
            data['percentage'] = 0
    
    return render_template('view_course_attendance.html', 
                         course=course, 
                         student_data=student_data)
@bp.route('/teacher/manage-enrollments/<int:course_id>', methods=['GET', 'POST'])
@login_required
def manage_enrollments(course_id):
    """Manage student enrollments in a course"""
    if current_user.role != 'teacher':
        flash('Access denied', 'danger')
        return redirect(url_for('main.index'))
    
    course = Course.query.get_or_404(course_id)
    
    if course.teacher_id != current_user.id:
        flash('Access denied', 'danger')
        return redirect(url_for('main.teacher_dashboard'))
    
    if request.method == 'POST':
        student_ids = request.form.getlist('students')
        
        # Remove all existing enrollments
        Enrollment.query.filter_by(course_id=course_id).delete()
        
        # Add new enrollments
        for student_id in student_ids:
            enrollment = Enrollment(
                student_id=int(student_id),
                course_id=course_id
            )
            db.session.add(enrollment)
        
        db.session.commit()
        flash(f'Enrolled {len(student_ids)} students in {course.name}', 'success')
        return redirect(url_for('main.teacher_dashboard'))
    
    # GET request - show form
    all_students = User.query.filter_by(role='student').all()
    enrolled_student_ids = [e.student_id for e in Enrollment.query.filter_by(course_id=course_id).all()]
    
    return render_template('manage_enrollments.html',
                         course=course,
                         all_students=all_students,
                         enrolled_student_ids=enrolled_student_ids)

@bp.route('/teacher/bulk-enroll/<int:course_id>', methods=['POST'])
@login_required
def bulk_enroll(course_id):
    """Bulk enroll all students in a course"""
    if current_user.role != 'teacher':
        flash('Access denied', 'danger')
        return redirect(url_for('main.index'))
    
    course = Course.query.get_or_404(course_id)
    
    if course.teacher_id != current_user.id:
        flash('Access denied', 'danger')
        return redirect(url_for('main.teacher_dashboard'))
    
    # Enroll all students
    all_students = User.query.filter_by(role='student').all()
    
    for student in all_students:
        # Check if already enrolled
        existing = Enrollment.query.filter_by(
            student_id=student.id,
            course_id=course_id
        ).first()
        
        if not existing:
            enrollment = Enrollment(
                student_id=student.id,
                course_id=course_id
            )
            db.session.add(enrollment)
    
    db.session.commit()
    flash(f'Enrolled all {len(all_students)} students', 'success')
    return redirect(url_for('main.manage_enrollments', course_id=course_id))
@bp.route('/teacher/export-attendance/<int:course_id>')
@login_required
def export_attendance(course_id):
    """Export attendance to Excel"""
    if current_user.role != 'teacher':
        flash('Access denied', 'danger')
        return redirect(url_for('main.index'))
    
    course = Course.query.get_or_404(course_id)
    
    if course.teacher_id != current_user.id:
        flash('Access denied', 'danger')
        return redirect(url_for('main.teacher_dashboard'))
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    # Add title
    ws['A1'] = f"Attendance Report - {course.name}"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = f"Course Code: {course.code}"
    ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    # Headers
    headers = ['Roll No', 'Student Name', 'Email', 'Total Classes', 'Present', 'Absent', 'Late', 'Attendance %', 'Status']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Get attendance data
    records = Attendance.query.filter_by(course_id=course_id).all()
    
    student_data = {}
    for record in records:
        if record.student_id not in student_data:
            student_data[record.student_id] = {
                'student': record.student,
                'total': 0,
                'present': 0,
                'absent': 0,
                'late': 0
            }
        student_data[record.student_id]['total'] += 1
        student_data[record.student_id][record.status] += 1
    
    # Add data
    row = 6
    for student_id, data in student_data.items():
        percentage = round((data['present'] / data['total']) * 100, 2) if data['total'] > 0 else 0
        status = 'Good' if percentage >= 75 else 'Low'
        
        ws.cell(row=row, column=1, value=data['student'].roll_no or 'N/A')
        ws.cell(row=row, column=2, value=data['student'].username)
        ws.cell(row=row, column=3, value=data['student'].email)
        ws.cell(row=row, column=4, value=data['total'])
        ws.cell(row=row, column=5, value=data['present'])
        ws.cell(row=row, column=6, value=data['absent'])
        ws.cell(row=row, column=7, value=data['late'])
        ws.cell(row=row, column=8, value=f"{percentage}%")
        ws.cell(row=row, column=9, value=status)
        
        # Color code status
        status_cell = ws.cell(row=row, column=9)
        if percentage >= 75:
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        row += 1
    
    # Adjust column widths
    for col in range(1, 10):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    # Save to BytesIO
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Send file
    filename = f"attendance_{course.code}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
@bp.route('/teacher/predictive-alerts/<int:course_id>')
@login_required
def predictive_alerts(course_id):
    """Show predictive attendance alerts"""
    if current_user.role != 'teacher':
        flash('Access denied', 'danger')
        return redirect(url_for('main.index'))
    
    course = Course.query.get_or_404(course_id)
    
    if course.teacher_id != current_user.id:
        flash('Access denied', 'danger')
        return redirect(url_for('main.teacher_dashboard'))
    
    # Get all enrolled students
    enrollments = Enrollment.query.filter_by(course_id=course_id).all()
    
    predictions = []
    for enrollment in enrollments:
        student = enrollment.student
        
        # Calculate current attendance
        records = Attendance.query.filter_by(
            student_id=student.id,
            course_id=course_id
        ).all()
        
        total = len(records)
        present = len([r for r in records if r.status == 'present'])
        
        if total == 0:
            continue
        
        current_percentage = (present / total) * 100
        
        # Predict: If student misses next 3 classes
        predicted_total = total + 3
        predicted_present = present
        predicted_percentage = (predicted_present / predicted_total) * 100
        
        # Calculate how many more classes they can miss
        min_required = 0.75  # 75%
        max_absences = int((total - (min_required * total)) / (1 - min_required))
        absences_so_far = total - present
        absences_remaining = max_absences - absences_so_far
        
        # Alert conditions
        alert_level = 'safe'
        if current_percentage < 75:
            alert_level = 'critical'
        elif predicted_percentage < 75:
            alert_level = 'warning'
        elif absences_remaining <= 2:
            alert_level = 'caution'
        
        predictions.append({
            'student': student,
            'current_percentage': round(current_percentage, 2),
            'predicted_percentage': round(predicted_percentage, 2),
            'absences_remaining': absences_remaining,
            'alert_level': alert_level,
            'total_classes': total,
            'present': present,
            'absent': total - present
        })
    
    # Sort by alert level
    alert_order = {'critical': 0, 'warning': 1, 'caution': 2, 'safe': 3}
    predictions.sort(key=lambda x: alert_order[x['alert_level']])
    
    return render_template('predictive_alerts.html',
                         course=course,
                         predictions=predictions)
@bp.route('/pricing')
def pricing():
    """Pricing page"""
    return render_template('pricing.html')
